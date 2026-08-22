import sys
from pathlib import Path
from openpyxl import load_workbook

VALID_STATUSES = {"draft", "scheduled", "live", "retired"}
REQUIRED_COLUMNS = [
    "questionId",
    "category",
    "subcategory",
    "status",
    "answer",
    "imagePath",
    "clue1",
    "clue2",
    "clue3",
    "clue4",
    "clue5",
    "clue6",
    "clue7",
    "clue8",
    "clue9",
    "clue10",
    "schemaVersion",
]

def fail(messages):
    print("\nVALIDATION FAILED")
    print("=" * 60)
    for message in messages:
        print(f"ERROR: {message}")
    sys.exit(1)

def main():
    base_dir = Path(__file__).resolve().parent

    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        candidates = sorted(base_dir.glob("*.xlsx"))
        if not candidates:
            fail(["No .xlsx file found in this folder."])
        if len(candidates) > 1:
            print("Multiple .xlsx files found:")
            for item in candidates:
                print(f" - {item.name}")
            fail(["Run the script with the workbook filename as an argument."])
        workbook_path = candidates[0]

    if not workbook_path.exists():
        fail([f"Workbook not found: {workbook_path}"])

    print(f"Reading: {workbook_path.name}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    if "Birds Import" in wb.sheetnames:
        ws = wb["Birds Import"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        fail(["Worksheet is empty."])

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    errors = []
    missing_columns = [c for c in REQUIRED_COLUMNS if c not in header_map]
    if missing_columns:
        fail([f"Missing required column(s): {', '.join(missing_columns)}"])

    records = []
    seen_ids = set()
    total_clues = 0

    for excel_row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def get(name):
            idx = header_map[name]
            if idx >= len(row):
                return ""
            value = row[idx]
            return "" if value is None else str(value).strip()

        question_id = get("questionId")
        category = get("category")
        subcategory = get("subcategory")
        status = get("status")
        answer = get("answer")
        image_path = get("imagePath")
        schema_version = get("schemaVersion")
        clues = [get(f"clue{i}") for i in range(1, 11)]

        if not question_id:
            errors.append(f"Row {excel_row_num}: questionId is blank.")
        elif question_id in seen_ids:
            errors.append(f"Row {excel_row_num}: duplicate questionId '{question_id}'.")
        else:
            seen_ids.add(question_id)

        if not category:
            errors.append(f"Row {excel_row_num}: category is blank.")
        if not subcategory:
            errors.append(f"Row {excel_row_num}: subcategory is blank.")
        if status not in VALID_STATUSES:
            errors.append(
                f"Row {excel_row_num}: invalid status '{status}'. "
                f"Use one of: {', '.join(sorted(VALID_STATUSES))}."
            )
        if not answer:
            errors.append(f"Row {excel_row_num}: answer is blank.")

        if not image_path:
            errors.append(f"Row {excel_row_num}: imagePath is blank.")
        else:
            if not image_path.startswith("challenge_images/"):
                errors.append(
                    f"Row {excel_row_num}: imagePath must start with 'challenge_images/'."
                )
            if not image_path.lower().endswith(".webp"):
                errors.append(
                    f"Row {excel_row_num}: imagePath must point to a .webp file."
                )

        missing_clues = [str(i) for i, clue in enumerate(clues, start=1) if not clue]
        if missing_clues:
            errors.append(
                f"Row {excel_row_num}: missing clue(s): {', '.join(missing_clues)}."
            )
        else:
            total_clues += 10

        if schema_version not in {"1", "1.0"}:
            errors.append(
                f"Row {excel_row_num}: schemaVersion should be 1, found '{schema_version}'."
            )

        records.append({
            "questionId": question_id,
            "category": category,
            "subcategory": subcategory,
            "status": status,
            "answer": answer,
            "imagePath": image_path,
            "clues": clues,
        })

    if not records:
        errors.append("No challenge rows found.")

    if errors:
        fail(errors)

    print("\nVALIDATION PASSED")
    print("=" * 60)
    print(f"Questions found: {len(records)}")
    print(f"Unique question IDs: {len(seen_ids)}")
    print(f"Clues found: {total_clues}")
    print("All questions have exactly 10 clues")
    print("All required answers are present")
    print("All image paths use challenge_images/... and end in .webp")
    print("All statuses are valid")
    print("schemaVersion is 1 for every row")
    print("\nNo Firebase data was changed.")
    print("This validator is read-only.")

if __name__ == "__main__":
    main()
