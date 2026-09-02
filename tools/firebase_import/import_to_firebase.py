import argparse
import re
import sys
from pathlib import Path

import firebase_admin
from firebase_admin import credentials, firestore, storage
from openpyxl import load_workbook

VALID_STATUSES = {"draft", "scheduled", "live", "retired"}
QUESTION_ID_PATTERN = re.compile(r"^.+_\d{4}$")
REQUIRED_COLUMNS = [
    "questionId",
    "category",
    "subcategory",
    "status",
    "answer",
    "imagePath",
    "clue1",
    "clue2",
    "clue3",
    "clue4",
    "clue5",
    "clue6",
    "clue7",
    "clue8",
    "clue9",
    "clue10",
    "schemaVersion",
]

FIREBASE_KEY_DIR = Path(r"C:\RhiPlay\firebase_keys")
STORAGE_BUCKET = "first-guess-a18b1.firebasestorage.app"
PROJECT_ROOT = Path(r"C:\RhiPlay\first_guess")
LOCAL_CATEGORY_IMAGE_ROOT = PROJECT_ROOT / "assets" / "images" / "categories"
COLLECTION_NAME = "challenges"


def stop(messages):
    print("\nIMPORT STOPPED")
    print("=" * 72)
    if isinstance(messages, str):
        messages = [messages]
    for message in messages:
        print(f"ERROR: {message}")
    sys.exit(1)


def find_single_file(folder: Path, pattern: str, description: str) -> Path:
    matches = sorted(folder.glob(pattern))
    if not matches:
        stop(f"No {description} found in {folder}")
    if len(matches) > 1:
        print(f"Multiple {description}s found:")
        for item in matches:
            print(f" - {item.name}")
        stop(f"Please specify which {description} to use.")
    return matches[0]


def find_workbook(base_dir: Path, workbook_arg: str | None) -> Path:
    if workbook_arg:
        path = Path(workbook_arg).expanduser()
        if not path.is_absolute():
            path = (base_dir / path).resolve()
        else:
            path = path.resolve()
        if not path.exists():
            stop(f"Workbook not found: {path}")
        return path
    return find_single_file(base_dir, "*.xlsx", "Excel workbook")


def find_service_account_key() -> Path:
    if not FIREBASE_KEY_DIR.exists():
        stop(f"Firebase key folder not found: {FIREBASE_KEY_DIR}")
    return find_single_file(
        FIREBASE_KEY_DIR, "*.json", "Firebase service-account JSON key"
    )


def split_accepted_answers(value) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    items = [item.strip().lower() for item in text.split("*")]
    return [item for item in items if item]


def read_and_validate(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    # Generic importer: use the first worksheet in any First Guess import workbook.
    # Category and subcategory come from each spreadsheet row, not the sheet name.
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        stop("Worksheet is empty.")

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    missing_columns = [c for c in REQUIRED_COLUMNS if c not in header_map]
    if missing_columns:
        stop(f"Missing required column(s): {', '.join(missing_columns)}")

    records = []
    errors = []
    seen_ids = set()

    for excel_row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def get(name):
            idx = header_map.get(name)
            if idx is None or idx >= len(row):
                return ""
            value = row[idx]
            return "" if value is None else str(value).strip()

        question_id = get("questionId")
        category = get("category").lower()
        subcategory = get("subcategory").lower()
        status = get("status").lower()
        answer = get("answer").lower()
        accepted_answers = split_accepted_answers(
            row[header_map["acceptedAnswers"]]
            if "acceptedAnswers" in header_map and header_map["acceptedAnswers"] < len(row)
            else None
        )
        image_path = get("imagePath").replace("\\", "/")
        clues = [get(f"clue{i}") for i in range(1, 11)]
        schema_raw = get("schemaVersion")

        if not question_id:
            errors.append(f"Row {excel_row_num}: questionId is blank.")
        elif not QUESTION_ID_PATTERN.fullmatch(question_id):
            errors.append(
                f"Row {excel_row_num}: questionId '{question_id}' is invalid. "
                f"Permanent Question IDs must end in exactly four digits, "
                f"for example 'animals_birds_0001'."
            )
        elif question_id in seen_ids:
            errors.append(f"Row {excel_row_num}: duplicate questionId '{question_id}'.")
        else:
            seen_ids.add(question_id)

        if not category:
            errors.append(f"Row {excel_row_num}: category is blank.")
        if not subcategory:
            errors.append(f"Row {excel_row_num}: subcategory is blank.")
        if status not in VALID_STATUSES:
            errors.append(
                f"Row {excel_row_num}: status '{status}' is invalid. "
                f"Use: {', '.join(sorted(VALID_STATUSES))}."
            )
        if not answer:
            errors.append(f"Row {excel_row_num}: answer is blank.")

        if not image_path:
            errors.append(f"Row {excel_row_num}: imagePath is blank.")
        else:
            if not image_path.startswith("challenge_images/"):
                errors.append(
                    f"Row {excel_row_num}: imagePath must start with 'challenge_images/'."
                )
            if not image_path.lower().endswith(".webp"):
                errors.append(
                    f"Row {excel_row_num}: imagePath must end in .webp."
                )

        missing_clues = [str(i) for i, clue in enumerate(clues, start=1) if not clue]
        if missing_clues:
            errors.append(
                f"Row {excel_row_num}: missing clue(s): {', '.join(missing_clues)}."
            )

        try:
            schema_version = int(float(schema_raw))
        except Exception:
            schema_version = None
        if schema_version != 1:
            errors.append(
                f"Row {excel_row_num}: schemaVersion must be 1; found '{schema_raw}'."
            )

        records.append(
            {
                "questionId": question_id,
                "category": category,
                "subcategory": subcategory,
                "status": status,
                "answer": answer,
                "acceptedAnswers": accepted_answers,
                "imagePath": image_path,
                "clues": clues,
                "schemaVersion": schema_version,
                "_excelRow": excel_row_num,
            }
        )

    if not records:
        errors.append("No challenge rows found.")

    if errors:
        stop(errors)

    return records


def resolve_local_image(record, replace_images=False):
    image_path = record["imagePath"].replace("\\", "/")
    prefix = "challenge_images/"

    if not image_path.startswith(prefix):
        stop(
            f"{record['questionId']}: imagePath does not start with '{prefix}': "
            f"{image_path}"
        )

    relative = image_path[len(prefix):]
    parts = Path(relative).parts

    if len(parts) < 3:
        stop(
            f"{record['questionId']}: imagePath must contain "
            f"<category>/<subcategory>/<filename>: {image_path}"
        )

    local_path = LOCAL_CATEGORY_IMAGE_ROOT.joinpath(*parts)

    if replace_images:
        optimized_path = local_path.parent / "optimized" / local_path.name
        if optimized_path.exists() and optimized_path.is_file():
            local_path = optimized_path

    if not local_path.exists():
        stop(
            f"{record['questionId']}: expected local image was not found:\n"
            f"  {local_path}\n"
            f"From Firebase path:\n"
            f"  {image_path}"
        )

    if not local_path.is_file():
        stop(f"{record['questionId']}: local image path is not a file: {local_path}")

    return local_path


def init_firebase(key_path: Path):
    if not firebase_admin._apps:
        cred = credentials.Certificate(str(key_path))
        firebase_admin.initialize_app(cred, {"storageBucket": STORAGE_BUCKET})
    return firestore.client(), storage.bucket()


def make_firestore_payload(record, existing: bool):
    payload = {
        "questionId": record["questionId"],
        "category": record["category"],
        "subcategory": record["subcategory"],
        "status": record["status"],
        "answer": record["answer"],
        "acceptedAnswers": record["acceptedAnswers"],
        "imagePath": record["imagePath"],
        "clues": record["clues"],
        "schemaVersion": record["schemaVersion"],
        "updatedAt": firestore.SERVER_TIMESTAMP,
    }
    if not existing:
        payload["createdAt"] = firestore.SERVER_TIMESTAMP
    return payload


def main():
    parser = argparse.ArgumentParser(
        description="First Guess bulk Firebase importer. Dry-run by default."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        help="Optional Excel workbook filename/path. If omitted, uses the only .xlsx in this folder.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Actually upload missing images and create/update Firestore documents.",
    )
    parser.add_argument(
        "--replace-images",
        action="store_true",
        help=(
            "Overwrite existing Storage images from the workbook's local image paths. "
            "Firestore is left untouched unless --write is also used without this mode."
        ),
    )
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    workbook_path = find_workbook(base_dir, args.workbook)
    key_path = find_service_account_key()

    print("FIRST GUESS FIREBASE BULK IMPORTER")
    print("=" * 72)
    if args.replace_images:
        mode = "REPLACE IMAGES" if args.write else "REPLACE IMAGES DRY RUN (NO CHANGES)"
    else:
        mode = "WRITE" if args.write else "DRY RUN (NO CHANGES)"
    print(f"Mode: {mode}")
    print(f"Workbook: {workbook_path.name}")
    print(f"Firestore collection: {COLLECTION_NAME}")
    print(f"Storage bucket: {STORAGE_BUCKET}")

    records = read_and_validate(workbook_path)
    print(f"\nSpreadsheet validation: PASSED ({len(records)} questions)")

    print("Checking exact local image paths...")
    local_images = {}
    for record in records:
        local_path = resolve_local_image(
            record,
            replace_images=args.replace_images,
        )
        local_images[record["questionId"]] = local_path
        print(f"LOCAL OK {record['questionId']} -> {local_path}")

    print(f"Local image validation: PASSED ({len(local_images)} images found)")

    db, bucket = init_firebase(key_path)

    print("\nChecking Firebase...")
    print("-" * 72)

    plan = []
    for record in records:
        qid = record["questionId"]
        blob = bucket.blob(record["imagePath"])
        remote_image_exists = blob.exists()

        doc_ref = db.collection(COLLECTION_NAME).document(qid)
        snapshot = doc_ref.get()
        document_exists = snapshot.exists

        if args.replace_images:
            image_action = "REPLACE" if remote_image_exists else "UPLOAD"
            doc_action = "KEEP"
        else:
            image_action = "KEEP existing" if remote_image_exists else "UPLOAD"
            doc_action = "UPDATE" if document_exists else "CREATE"

        print(
            f"{qid}: image={image_action:<13} firestore={doc_action:<6} "
            f"answer='{record['answer']}'"
        )

        plan.append(
            {
                "record": record,
                "local_image": local_images[qid],
                "remote_image_exists": remote_image_exists,
                "document_exists": document_exists,
                "doc_ref": doc_ref,
                "blob": blob,
            }
        )

    print("-" * 72)

    upload_count = sum(not item["remote_image_exists"] for item in plan)
    replace_count = sum(item["remote_image_exists"] for item in plan) if args.replace_images else 0
    create_count = 0 if args.replace_images else sum(not item["document_exists"] for item in plan)
    update_count = 0 if args.replace_images else sum(item["document_exists"] for item in plan)

    print("\nIMPORT PLAN")
    print(f"Questions: {len(plan)}")
    if args.replace_images:
        print(f"Existing Storage images to replace: {replace_count}")
        print(f"Missing Storage images to upload: {upload_count}")
        print("Firestore documents to create: 0")
        print("Firestore documents to update: 0")
    else:
        print(f"Images already in Storage: {len(plan) - upload_count}")
        print(f"Images to upload: {upload_count}")
        print(f"Firestore documents to create: {create_count}")
        print(f"Firestore documents to update: {update_count}")

    if not args.write:
        print("\nDRY RUN COMPLETE — NO FIREBASE DATA WAS CHANGED.")
        print("If the plan above is correct, run:")
        if args.replace_images:
            print(
                f'    python import_to_firebase.py "{workbook_path.name}" '
                "--replace-images --write"
            )
        else:
            print(f'    python import_to_firebase.py "{workbook_path.name}" --write')
        return

    print("\nWRITE MODE STARTING")
    print("=" * 72)

    uploaded = 0
    replaced = 0
    for item in plan:
        if item["remote_image_exists"] and not args.replace_images:
            continue

        local_path = item["local_image"]
        blob = item["blob"]
        was_existing = item["remote_image_exists"]
        blob.upload_from_filename(
            str(local_path),
            content_type="image/webp",
        )

        if was_existing:
            replaced += 1
            print(
                f"REPLACED IMAGE: {item['record']['questionId']} -> "
                f"{item['record']['imagePath']}"
            )
        else:
            uploaded += 1
            print(
                f"UPLOADED IMAGE: {item['record']['questionId']} -> "
                f"{item['record']['imagePath']}"
            )

    if args.replace_images:
        print("\nIMAGE REPLACEMENT COMPLETE")
        print("=" * 72)
        print(f"Images replaced: {replaced}")
        print(f"Missing images uploaded: {uploaded}")
        print("Firestore documents written: 0")
        print("Permanent question IDs and Firestore data were left untouched.")
        return

    batch = db.batch()
    write_ops = 0
    committed = 0

    for item in plan:
        record = item["record"]
        payload = make_firestore_payload(record, item["document_exists"])
        batch.set(item["doc_ref"], payload, merge=True)
        write_ops += 1

        if write_ops == 400:
            batch.commit()
            committed += write_ops
            batch = db.batch()
            write_ops = 0

    if write_ops:
        batch.commit()
        committed += write_ops

    print("\nIMPORT COMPLETE")
    print("=" * 72)
    print(f"Images uploaded: {uploaded}")
    print(f"Firestore documents written: {committed}")
    print(f"Created: {create_count}")
    print(f"Updated: {update_count}")
    print("Existing Storage images were left untouched.")
    print("Permanent question IDs were used as Firestore document IDs.")


if __name__ == "__main__":
    main()
